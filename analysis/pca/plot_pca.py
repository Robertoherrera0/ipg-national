import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


df = pd.read_excel("IPG_metrics_updated.xlsx")
school_col = df.columns[0]
df = df.dropna(axis=1, how="all")

full_metrics = [
    "EdgesPerNode",
    "assortativity_unweighted",
    "assortativity_weighted",
    "clique_integration",
    "avg_path_over_diameter",
    "global_efficiency",
    "local_efficiency",
    "sigma_small_world_index",
    "omega_small_world_index"
]

full_metrics = [m for m in full_metrics if m in df.columns]
df_fullM = df[full_metrics].apply(pd.to_numeric, errors="coerce")
mask_full = df_fullM.notna().all(axis=1)
df_full = df.loc[mask_full].reset_index(drop=True)
df_fullM = df_full[full_metrics]

full_outliers = ["Maine"]
mask_full_clean = ~df_full[school_col].isin(full_outliers)
df_full_clean = df_full[mask_full_clean].reset_index(drop=True)
df_full_cleanM = df_full_clean[full_metrics]

X_full = StandardScaler().fit_transform(df_fullM)
pca_full = PCA(n_components=2)
PC_full = pca_full.fit_transform(X_full)
df_full["PC1"] = PC_full[:,0]
df_full["PC2"] = PC_full[:,1]

X_full_k = StandardScaler().fit_transform(df_full_cleanM)
labels_full = KMeans(n_clusters=3, random_state=42, n_init=30).fit_predict(X_full_k) + 1

df_full["Cluster"] = 0
df_full.loc[mask_full_clean, "Cluster"] = labels_full

combined_metrics = [""]

combined_metrics = [m for m in combined_metrics if m in df.columns]
df_combM = df[combined_metrics].apply(pd.to_numeric, errors="coerce")
mask_comb = df_combM.notna().all(axis=1)
df_comb = df.loc[mask_comb].reset_index(drop=True)
df_combM = df_comb[combined_metrics]

filtered_outliers = ["Delaware", "Wyoming", "Maine"]
mask_comb_clean = ~df_comb[school_col].isin(filtered_outliers)
df_comb_clean = df_comb[mask_comb_clean].reset_index(drop=True)
df_comb_cleanM = df_comb_clean[combined_metrics]

X_comb = StandardScaler().fit_transform(df_combM)
pca_comb = PCA(n_components=2)
PC_comb = pca_comb.fit_transform(X_comb)
df_comb["PC1"] = PC_comb[:,0]
df_comb["PC2"] = PC_comb[:,1]

X_comb_k = StandardScaler().fit_transform(df_comb_cleanM)
labels_comb = KMeans(n_clusters=3, random_state=42, n_init=30).fit_predict(X_comb_k) + 1

df_comb["Cluster"] = 0
df_comb.loc[mask_comb_clean, "Cluster"] = labels_comb

label_fontsize = 5
title_pad = 12
offset = 0.01

fig, ax = plt.subplots(2,2, figsize=(18,14))

ax0 = ax[0,0]
for c in sorted(df_full["Cluster"].unique()):
    sub = df_full[df_full["Cluster"]==c]
    color = "black" if c == 0 else None
    size = 100 if c == 0 else 60
    ax0.scatter(sub["PC1"], sub["PC2"], s=size, c=color, label=("Outliers" if c==0 else f"Cluster {c}"))
    for _,r in sub.iterrows():
        ax0.text(r["PC1"]+offset, r["PC2"]+offset, r[school_col], fontsize=label_fontsize)
ax0.set_title("PCA — Full Graph Metrics (K=3)", pad=title_pad)
ax0.set_xlabel("PC1")
ax0.set_ylabel("PC2")
ax0.grid(True)
ax0.legend()

load_full = pd.DataFrame(pca_full.components_.T, index=full_metrics, columns=["PC1","PC2"])
ax1 = ax[0,1]
x = np.arange(len(full_metrics))
bw = 0.35
ax1.bar(x - bw/2, load_full["PC1"], width=bw, label="PC1")
ax1.bar(x + bw/2, load_full["PC2"], width=bw, label="PC2")
ax1.set_xticks(x)
ax1.set_xticklabels(full_metrics, rotation=75, ha='right', fontsize=7)
ax1.set_title("Loadings — Full Graph Metrics", pad=title_pad)
ax1.grid(True, axis="y")
ax1.legend()

ax2 = ax[1,0]
for c in sorted(df_comb["Cluster"].unique()):
    sub = df_comb[df_comb["Cluster"]==c]
    color = "black" if c == 0 else None
    size = 100 if c == 0 else 60
    ax2.scatter(sub["PC1"], sub["PC2"], s=size, c=color, label=("Outliers" if c==0 else f"Cluster {c}"))
    for _,r in sub.iterrows():
        ax2.text(r["PC1"]+offset, r["PC2"]+offset, r[school_col], fontsize=label_fontsize)
ax2.set_title("PCA — Filtered Metrics (K=3)", pad=title_pad)
ax2.set_xlabel("PC1")
ax2.set_ylabel("PC2")
ax2.grid(True)
ax2.legend()

load_comb = pd.DataFrame(pca_comb.components_.T, index=combined_metrics, columns=["PC1","PC2"])
ax3 = ax[1,1]
x2 = np.arange(len(combined_metrics))
ax3.bar(x2 - bw/2, load_comb["PC1"], width=bw, label="PC1")
ax3.bar(x2 + bw/2, load_comb["PC2"], width=bw, label="PC2")
ax3.set_xticks(x2)
ax3.set_xticklabels(combined_metrics, rotation=75, ha='right', fontsize=7)
ax3.set_title("Loadings — Filtered Metrics", pad=title_pad)
ax3.grid(True, axis="y")
ax3.legend()

plt.tight_layout()
plt.savefig("IPG_groups.png", dpi=400, bbox_inches="tight")
plt.show()
wb = Workbook()

def write_block(ws, title, df, metrics, start_row):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    start_row += 1
    for col_idx, col_name in enumerate([school_col] + metrics, 1):
        c = ws.cell(row=start_row, column=col_idx, value=col_name)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
    start_row += 1
    for _, r in df.iterrows():
        for col_idx, col_name in enumerate([school_col] + metrics, 1):
            ws.cell(row=start_row, column=col_idx, value=r[col_name])
        start_row += 1
    return start_row + 1

ws1 = wb.active
ws1.title = "Full Graph"

row = 1
for c in [1,2,3]:
    block = df_full[df_full["Cluster"]==c]
    row = write_block(ws1, f"Cluster {c}", block, full_metrics, row)

out_block = df_full[df_full["Cluster"]==0]
row = write_block(ws1, "Outliers", out_block, full_metrics, row)

missing_full = df[~df[school_col].isin(df_full[school_col])]
row = write_block(ws1, "Dropped - Missing Data", missing_full, full_metrics, row)

for col in ws1.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
    ws1.column_dimensions[col[0].column_letter].width = max_len + 2

ws2 = wb.create_sheet("Removed Weight = 1")

row = 1
for c in [1,2,3]:
    block = df_comb[df_comb["Cluster"]==c]
    row = write_block(ws2, f"Cluster {c}", block, combined_metrics, row)

out_block_f = df_comb[df_comb["Cluster"]==0]
row = write_block(ws2, "Outliers", out_block_f, combined_metrics, row)

missing_comb = df[~df[school_col].isin(df_comb[school_col])]
row = write_block(ws2, "Dropped - Missing Data", missing_comb, combined_metrics, row)

for col in ws2.columns:
    max_len = max(len(str(c.value)) if c.value else 0 for c in col)
    ws2.column_dimensions[col[0].column_letter].width = max_len + 2

wb.save("IPG_groups.xlsx")
